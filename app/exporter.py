"""Export messages to various formats."""
import csv
import json
import os
from datetime import datetime
from typing import Optional

from app.models import MessageData


class Exporter:
    """Exports messages to JSON, CSV, Excel, HTML, PDF."""

    def __init__(self, messages: list[MessageData], chat_title: str = "Chat"):
        self.messages = messages
        self.chat_title = chat_title

    def to_json(self, path: str) -> str:
        """Export to JSON."""
        data = {
            "chat": self.chat_title,
            "exported_at": datetime.now().isoformat(),
            "total_messages": len(self.messages),
            "messages": [m.to_dict() for m in self.messages],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return path

    def to_csv(self, path: str) -> str:
        """Export to CSV."""
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Date", "Sender ID", "Sender", "Text", "Media", "Reply To", "Views"])
            for m in self.messages:
                writer.writerow([
                    m.id, m.date.strftime("%Y-%m-%d %H:%M:%S"),
                    m.sender_id, m.sender_name, m.text,
                    m.media_type or "", m.reply_to_id or "", m.views or "",
                ])
        return path

    def to_excel(self, path: str) -> str:
        """Export to Excel (.xlsx)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["ID", "Date", "Sender ID", "Sender", "Text", "Media", "Reply To", "Views"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, m in enumerate(self.messages, 2):
            values = [
                m.id, m.date.strftime("%Y-%m-%d %H:%M:%S"),
                m.sender_id, m.sender_name, m.text,
                m.media_type or "", m.reply_to_id or "", m.views or "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col == 5:
                    cell.alignment = Alignment(wrap_text=True)

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 10

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Chat"
        ws2["B1"] = self.chat_title
        ws2["A2"] = "Total Messages"
        ws2["B2"] = len(self.messages)
        ws2["A3"] = "Export Date"
        ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws2["A4"] = "Date Range"
        if self.messages:
            ws2["B4"] = f"{self.messages[0].date.strftime('%Y-%m-%d')} - {self.messages[-1].date.strftime('%Y-%m-%d')}"

        # Sender stats
        ws2["A6"] = "Sender Statistics"
        ws2["A6"].font = Font(bold=True, size=12)
        ws2["A7"] = "Sender"
        ws2["B7"] = "Messages"
        ws2["A7"].font = header_font
        ws2["A7"].fill = header_fill
        ws2["B7"].font = header_font
        ws2["B7"].fill = header_fill

        sender_stats = {}
        for m in self.messages:
            sender_stats[m.sender_name] = sender_stats.get(m.sender_name, 0) + 1
        for i, (name, count) in enumerate(sorted(sender_stats.items(), key=lambda x: -x[1]), 8):
            ws2[f"A{i}"] = name
            ws2[f"B{i}"] = count

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 30

        wb.save(path)
        return path

    def to_html(self, path: str) -> str:
        """Export to HTML with beautiful styling."""
        from jinja2 import Template

        template = Template(HTML_TEMPLATE)
        html = template.render(
            chat_title=self.chat_title,
            messages=self.messages,
            total=len(self.messages),
            export_date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        return path

    def to_pdf(self, path: str) -> str:
        """Export to PDF."""
        from fpdf import FPDF

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Add Unicode font
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, f"Chat: {self._safe_latin(self.chat_title)}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, f"Total messages: {len(self.messages)}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        pdf.set_font("Helvetica", "", 9)
        for m in self.messages:
            date_str = m.date.strftime("%Y-%m-%d %H:%M")
            sender = self._safe_latin(m.sender_name)
            text = self._safe_latin(m.text[:500]) if m.text else "[no text]"

            header = f"[{date_str}] {sender}:"
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, header, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 5, text)
            if m.media_type:
                pdf.set_font("Helvetica", "I", 8)
                pdf.cell(0, 5, f"[{m.media_type}]", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

        pdf.output(path)
        return path

    @staticmethod
    def _safe_latin(text: str) -> str:
        """Replace non-latin chars for basic PDF (no Unicode font)."""
        return text.encode("latin-1", errors="replace").decode("latin-1")


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ chat_title }} - Message Export</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #0f0f0f; color: #e0e0e0; padding: 20px;
  }
  .header {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    border-radius: 16px; padding: 30px; margin-bottom: 20px;
    border: 1px solid #2a2a4a;
  }
  .header h1 { color: #7c8cf8; font-size: 24px; margin-bottom: 8px; }
  .header .meta { color: #888; font-size: 14px; }
  .header .meta span { color: #7c8cf8; font-weight: 600; }
  .messages { max-width: 900px; margin: 0 auto; }
  .msg {
    background: #1a1a2e; border-radius: 12px; padding: 16px; margin-bottom: 8px;
    border: 1px solid #2a2a3a; transition: border-color 0.2s;
  }
  .msg:hover { border-color: #7c8cf8; }
  .msg-header { display: flex; justify-content: space-between; margin-bottom: 6px; }
  .msg-sender { color: #7c8cf8; font-weight: 600; font-size: 14px; }
  .msg-date { color: #666; font-size: 12px; }
  .msg-text { color: #d0d0d0; font-size: 14px; line-height: 1.5; white-space: pre-wrap; word-break: break-word; }
  .msg-media { color: #f0a040; font-size: 12px; margin-top: 4px; font-style: italic; }
  .stats {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
    gap: 12px; margin-bottom: 20px;
  }
  .stat-card {
    background: #1a1a2e; border-radius: 12px; padding: 20px; text-align: center;
    border: 1px solid #2a2a4a;
  }
  .stat-card .value { color: #7c8cf8; font-size: 28px; font-weight: 700; }
  .stat-card .label { color: #888; font-size: 12px; margin-top: 4px; }
</style>
</head>
<body>
<div class="messages">
  <div class="header">
    <h1>{{ chat_title }}</h1>
    <div class="meta">
      Exported <span>{{ total }}</span> messages on <span>{{ export_date }}</span>
    </div>
  </div>
  <div class="stats">
    <div class="stat-card">
      <div class="value">{{ total }}</div>
      <div class="label">Messages</div>
    </div>
    <div class="stat-card">
      <div class="value">{{ messages | map(attribute='sender_name') | unique | list | length }}</div>
      <div class="label">Senders</div>
    </div>
    {% if messages %}
    <div class="stat-card">
      <div class="value">{{ messages[0].date.strftime('%m/%d') }}</div>
      <div class="label">First Message</div>
    </div>
    <div class="stat-card">
      <div class="value">{{ messages[-1].date.strftime('%m/%d') }}</div>
      <div class="label">Last Message</div>
    </div>
    {% endif %}
  </div>
  {% for m in messages %}
  <div class="msg">
    <div class="msg-header">
      <span class="msg-sender">{{ m.sender_name }}</span>
      <span class="msg-date">{{ m.date.strftime('%Y-%m-%d %H:%M:%S') }}</span>
    </div>
    {% if m.text %}<div class="msg-text">{{ m.text }}</div>{% endif %}
    {% if m.media_type %}<div class="msg-media">[{{ m.media_type }}]</div>{% endif %}
  </div>
  {% endfor %}
</div>
</body>
</html>"""
